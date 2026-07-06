import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

import models
import schemas
from database import get_db
from security import get_current_user_id

router = APIRouter(prefix="/api/v1/transactions/import", tags=["import"])

MAX_ROWS = 500
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
ALLOWED_EXTENSIONS = (".csv", ".xlsx")

DATE_FORMATS = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"]

COLUMN_ALIASES = {
    "date": ["tarih", "date"],
    "amount": ["tutar", "amount", "miktar"],
    "description": ["açıklama", "aciklama", "description", "not"],
    "type": ["tür", "tur", "type"],
    "category": ["kategori", "category"],
}


def _parse_date(value: str) -> date | None:
    value = value.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: str) -> float | None:
    value = value.strip().replace(" ", "").replace("₺", "").replace("TL", "")
    if not value:
        return None
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return None


def _find_column(headers: list[str], keys: list[str]) -> int | None:
    normalized = [h.strip().lower() for h in headers]
    for key in keys:
        if key in normalized:
            return normalized.index(key)
    return None


def _rows_from_csv(raw: bytes) -> list[list[str]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Dosya UTF-8 formatında olmalı")
    return [row for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(raw: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")

    sheet = workbook.active
    return [
        ["" if cell is None else str(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]


class ImportRow(BaseModel):
    row_number: int
    amount: float | None
    type: str
    category_id: int | None
    category_name_hint: str | None
    description: str | None
    date: date | None
    error: str | None


@router.post("/preview", response_model=list[ImportRow])
async def preview_import(
    file: UploadFile,
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user_id),
):
    filename = (file.filename or "").lower()
    if not filename.endswith(ALLOWED_EXTENSIONS):
        raise HTTPException(status_code=400, detail="Sadece .csv veya .xlsx dosyaları desteklenir")

    raw = await file.read()
    if len(raw) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Dosya çok büyük (maksimum 5 MB)")

    rows = _rows_from_xlsx(raw) if filename.endswith(".xlsx") else _rows_from_csv(raw)
    rows = [row for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        raise HTTPException(status_code=400, detail="Dosya boş")

    headers = rows[0]
    date_idx = _find_column(headers, COLUMN_ALIASES["date"])
    amount_idx = _find_column(headers, COLUMN_ALIASES["amount"])
    desc_idx = _find_column(headers, COLUMN_ALIASES["description"])
    type_idx = _find_column(headers, COLUMN_ALIASES["type"])
    category_idx = _find_column(headers, COLUMN_ALIASES["category"])

    if date_idx is None or amount_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Dosyada 'Tarih' ve 'Tutar' sütunları bulunamadı. Lütfen sütun başlıklarını kontrol et.",
        )

    user_categories = {
        c.name.strip().lower(): c.id
        for c in db.query(models.Category).filter(models.Category.user_id == user_id)
    }

    data_rows = rows[1 : 1 + MAX_ROWS]
    results: list[ImportRow] = []
    for i, row in enumerate(data_rows, start=2):
        parsed_date = _parse_date(row[date_idx]) if date_idx < len(row) else None
        parsed_amount = _parse_amount(row[amount_idx]) if amount_idx < len(row) else None
        description = (
            row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else None
        )
        category_name = (
            row[category_idx].strip()
            if category_idx is not None and category_idx < len(row) and row[category_idx].strip()
            else None
        )

        error = None
        if parsed_date is None:
            error = "Tarih okunamadı"
        elif parsed_amount is None:
            error = "Tutar okunamadı"

        if type_idx is not None and type_idx < len(row) and row[type_idx].strip():
            tx_type = "income" if row[type_idx].strip().lower() in ("gelir", "income") else "expense"
        else:
            tx_type = "income" if (parsed_amount or 0) > 0 else "expense"

        category_id = user_categories.get(category_name.lower()) if category_name else None

        results.append(
            ImportRow(
                row_number=i,
                amount=abs(parsed_amount) if parsed_amount is not None else None,
                type=tx_type,
                category_id=category_id,
                category_name_hint=category_name,
                description=description or None,
                date=parsed_date,
                error=error,
            )
        )

    return results


class ImportConfirmPayload(BaseModel):
    rows: list[schemas.TransactionCreate]


@router.post("/confirm")
def confirm_import(
    payload: ImportConfirmPayload,
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user_id),
):
    if len(payload.rows) == 0:
        raise HTTPException(status_code=400, detail="Aktarılacak satır yok")
    if len(payload.rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"En fazla {MAX_ROWS} satır aktarılabilir")

    category_ids = {r.category_id for r in payload.rows if r.category_id is not None}
    if category_ids:
        owned = {
            row.id
            for row in db.query(models.Category.id).filter(
                models.Category.user_id == user_id, models.Category.id.in_(category_ids)
            )
        }
        if category_ids - owned:
            raise HTTPException(status_code=404, detail="Geçersiz kategori")

    transactions = [
        models.Transaction(user_id=user_id, **row.model_dump()) for row in payload.rows
    ]
    db.add_all(transactions)
    db.commit()

    return {"imported": len(transactions)}
